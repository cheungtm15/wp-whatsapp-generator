import streamlit as st
import openpyxl
import urllib.parse
import io
import pandas as pd

# Page Configuration
st.set_page_config(page_title="WP WhatsApp Link Generator", page_icon="💬", layout="wide")

st.title("💬 WP WhatsApp Link Generator")
st.write("Upload **New Orders.xlsx** or **Today delivery.xlsx** to append a column with formatted, clickable WhatsApp links.")

# 1. User Choice for Template Type
message_type = st.radio(
    "Select the type of notification message you want to generate:",
    [
        "🚚 Today's Delivery Reminder (Uses Order No.)",
        "📅 Future Delivery Confirmation (Uses ETD / Delivery Date)"
    ]
)

# File Uploader
uploaded_file = st.file_uploader("Choose an Excel file (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    file_bytes = uploaded_file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        
        # Detect the target sheet (prefers 'Summary Data')
        if "Summary Data" in wb.sheetnames:
            sheet_name = "Summary Data"
        else:
            sheet_name = wb.sheetnames[0]
            for name in wb.sheetnames:
                ws = wb[name]
                headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                if "Sales Order (Remarks) Tel" in headers and "Sales Order Cus. P.O. No." in headers:
                    sheet_name = name
                    break
                    
        ws = wb[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        
        # Verify required base columns
        if "Sales Order (Remarks) Tel" not in headers or "Sales Order Cus. P.O. No." not in headers:
            st.error(f"⚠️ Could not find required columns ('Sales Order (Remarks) Tel' and 'Sales Order Cus. P.O. No.') in '{sheet_name}'.")
        else:
            tel_idx = headers.index("Sales Order (Remarks) Tel") + 1
            po_idx = headers.index("Sales Order Cus. P.O. No.") + 1
            
            # Additional column needed for Template 2 (Delivery Confirmation)
            etd_idx = headers.index("Sales Order ETD") + 1 if "Sales Order ETD" in headers else None
            
            if "📅 Future Delivery" in message_type and etd_idx is None:
                st.warning("⚠️ 'Sales Order ETD' column not found. The app will attempt to format links, but dates might fall back to 'today'.")

            # Determine or create the WhatsApp Link column index
            if "WhatsApp Link" in headers:
                link_col_idx = headers.index("WhatsApp Link") + 1
            else:
                link_col_idx = len(headers) + 1
                ws.cell(row=1, column=link_col_idx, value="WhatsApp Link")
                
            count = 0
            
            # Loop through rows to generate messages
            for row_idx in range(2, ws.max_row + 1):
                tel_val = ws.cell(row=row_idx, column=tel_idx).value
                po_val = ws.cell(row=row_idx, column=po_idx).value
                etd_val = ws.cell(row=row_idx, column=etd_idx).value if etd_idx else None
                
                if tel_val is not None:
                    # Clean & Format Phone Number
                    p_str = str(tel_val).strip()
                    if p_str.endswith('.0'): 
                        p_str = p_str[:-2]
                    p_str = "".join(c for c in p_str if c.isdigit())
                    if len(p_str) == 8:
                        p_str = "852" + p_str  # Default HK prefix for 8-digit numbers
                        
                    # Format Order Number
                    po_str = str(po_val).strip() if po_val else "your order"
                    
                    # Format Delivery Date nicely if available (e.g., "04 June 2026")
                    date_str = "today"
                    if etd_val:
                        try:
                            # If it's a datetime object or a string convert it
                            date_obj = pd.to_datetime(etd_val)
                            date_str = date_obj.strftime("%d %B %Y")
                        except:
                            date_str = str(etd_val).strip()

                    # Construct message based on user selection
                    if "🚚 Today's Delivery" in message_type:
                        # Template 1
                        text = (
                            f"Hello, Thank you for your order with WP at home\n"
                            f"Please be informed that we will deliver your order {po_str} today, between 11:00 am - 6:00 pm.\xa0\n"
                            f"Thank you and enjoy"
                        )
                    else:
                        # Template 2
                        text = (
                            f"Dear Customer,\n\xa0\n"
                            f"Thank you for your order!\n\n"
                            f"Your requested delivery date has been confirmed.\n\n"
                            f"We will deliver your order on\xa0 {date_str}, between 11:00 am - 6:00 pm.\n\n"
                            f"Please make sure the phone number provided is correct and have someone can access the door.\n\n"
                            f"Thank you and enjoy!\n\n"
                            f"WP at Home\n"
                            f"Phone: +852 2880 0000\n"
                            f"Email: wpathome@wavespacific.com"
                        )
                    
                    # URL Encode text and construct finalized API Link
                    encoded_text = urllib.parse.quote(text)
                    whatsapp_url = f"https://api.whatsapp.com/send?phone={p_str}&text={encoded_text}"
                    
                    # Apply to Excel Cell
                    cell = ws.cell(row=row_idx, column=link_col_idx, value="Send WhatsApp")
                    cell.hyperlink = whatsapp_url
                    cell.font = openpyxl.styles.Font(color="0563C1", underline="single")
                    count += 1
            
            # Save the modified spreadsheet out to buffer
            out_buffer = io.BytesIO()
            wb.save(out_buffer)
            out_buffer.seek(0)
            
            # Create a dataframe for previewing inside the browser app
            data = []
            updated_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
            for row_idx in range(2, ws.max_row + 1):
                row_vals = [ws.cell(row=row_idx, column=c_idx).value for c_idx in range(1, ws.max_column + 1)]
                if len(row_vals) >= link_col_idx:
                    h_target = ws.cell(row=row_idx, column=link_col_idx).hyperlink
                    row_vals[link_col_idx - 1] = h_target.target if h_target else ""
                data.append(row_vals)
                
            preview_df = pd.DataFrame(data, columns=updated_headers)
            
            st.success(f"🎉 Processed {count} records successfully using selected template!")
            
            # Download Button
            st.download_button(
                label="📥 Download Updated Spreadsheet",
                data=out_buffer,
                file_name=f"processed_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Interactive App Preview Dashboard
            st.subheader("👀 Generated Data Preview")
            st.dataframe(
                preview_df,
                column_config={
                    "WhatsApp Link": st.column_config.LinkColumn("WhatsApp Link", display_text="Test Link")
                },
                use_container_width=True
            )
            
    except Exception as e:
        st.error(f"❌ An error occurred while parsing the file: {e}")
